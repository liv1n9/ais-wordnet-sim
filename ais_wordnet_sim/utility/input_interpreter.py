import csv
import os

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


'''
    Lấy dữ liệu gồm:
        - câu hỏi 
        - câu trả lời
        - topic
        ... ở dạng text hoặc aiml
    Từ các nguồn:
        - csv 
        - xcel 
        - nguồn khác ... (web)

    Output dạng [(tuple1),(tuple2)], trong đó:
        tuple = (topic_name,text_question,text_answer)
'''

class InputInterpreter():
    FORMAT = ['csv','xlsx']
    def __init__(self):
        pass
    
    def validate_cell(self,input_cell):
        if not input_cell or input_cell == '' or input_cell == 'None':
            return False
        return True
    
    def interpret_csv(self,source_file_location,dest,
                        text_question_field,answer_field,topic_field):
        current_list = []
        with open(source_file_location,"r") as csv_file:
            readCSV = csv.DictReader(csv_file)            
            current_tuple = ()
            for row in readCSV:
                current_tuple = [row[topic_field],
                                row[text_question_field].replace("?","").strip(),
                                row[answer_field].strip()]
                current_list.append(current_tuple)
        
        return current_list

    def interpret_xlsx(self,source_file_location,dest,
                        text_question_field_as_col_num,
                        answer_field_as_col_num,
                        topic_field):
        qa_as_dict = {}
        workbook = load_workbook(filename=os.path.join(source_file_location))
        for each_sheet in workbook.sheetnames:
            worksheet = workbook[each_sheet]
            for row in range(1,worksheet.max_row):
                if self.validate_cell(str(worksheet.cell(column=1,row=row).value)) and self.validate_cell(str(worksheet.cell(column=2,row=row).value)):
                    qa_as_dict[str(worksheet.cell(column=1,row=row).value).replace("?","").strip()] = \
                str(worksheet.cell(column=2,row=row).value).strip()

        return qa_as_dict


